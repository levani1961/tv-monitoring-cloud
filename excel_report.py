from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from time_utils import add_seconds_to_datetime, seconds_to_hhmmss


REPORTS_DIR = Path("reports")


def build_excel_report(violations, output_path=None, metadata=None):
    REPORTS_DIR.mkdir(exist_ok=True)
    output_path = Path(output_path or REPORTS_DIR / "enterprise_hidden_ad_report.xlsx")
    metadata = metadata or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Hidden Ads"

    headers = [
        "Channel",
        "Broadcast Date",
        "Start Time (hh:mm:ss)",
        "End Time (hh:mm:ss)",
        "Genre",
        "Georgian Transcript",
        "Brand Name",
        "Probability Score (%)",
        "AI Risk Status",
        "AI Legal Reason",
        "Screenshot",
        "Video Proof",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    if not violations:
        row = [
            metadata.get("channel", "-"),
            metadata.get("date_range", "-"),
            "-",
            "-",
            metadata.get("genre", "-"),
            "დარღვევა ვერ მოიძებნა",
            "-",
            0,
            "No Violation Found",
            metadata.get("ignored_reason", ""),
            "-",
            "-",
        ]
        ws.append(row)
    else:
        for violation in violations:
            ws.append(
                [
                    violation.channel,
                    violation.broadcast_date,
                    seconds_to_hhmmss(violation.start),
                    seconds_to_hhmmss(violation.end),
                    violation.genre,
                    violation.transcript,
                    violation.brand_name,
                    violation.probability_score,
                    violation.risk_status,
                    violation.reason,
                    "",
                    "Open video proof",
                ]
            )

            row_index = ws.max_row
            if violation.screenshot_path and Path(violation.screenshot_path).exists():
                img = ExcelImage(violation.screenshot_path)
                img.width = 180
                img.height = 100
                ws.add_image(img, f"K{row_index}")
                ws.row_dimensions[row_index].height = 80

            if violation.clip_path:
                clip_cell = ws.cell(row=row_index, column=12)
                clip_cell.hyperlink = str(Path(violation.clip_path).resolve())
                clip_cell.style = "Hyperlink"

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    widths = {
        "A": 20,
        "B": 18,
        "C": 18,
        "D": 18,
        "E": 26,
        "F": 70,
        "G": 28,
        "H": 22,
        "I": 20,
        "J": 55,
        "K": 28,
        "L": 34,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)
    return output_path


def violation_rows_for_dashboard(violations):
    rows = []
    for violation in violations:
        rows.append(
            {
                "Channel": violation.channel,
                "Date": violation.broadcast_date,
                "Start": seconds_to_hhmmss(violation.start),
                "End": seconds_to_hhmmss(violation.end),
                "Genre": violation.genre,
                "Probability (%)": violation.probability_score,
                "Risk Level": violation.risk_status,
                "Brand": violation.brand_name,
                "Transcript": violation.transcript,
                "Clip": violation.clip_path,
            }
        )
    return rows
